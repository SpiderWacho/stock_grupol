import requests
import pandas as pd
import sqlite3
import win32com.client
from dotenv import load_dotenv
import warnings
import os
from openpyxl import load_workbook
from datetime import date

TARGET_EXCEL = r"C:\Users\GastonVecchio\Documents\Code\Python\Stocks\Stock depositos.xlsx"
PRODUCT_COL   = "Familia"               # Column name for product names
STOCK_COL     = "Cobertura" 

load_dotenv()
EMAIL_RECEIVER = os.getenv("EMAIL")

today = date.today()

conn = sqlite3.connect(r"C:\Users\GastonVecchio\Documents\Code\inventory.db")
cursor = conn.cursor()


def _migrate_db():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS stockouts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_name TEXT,
            date_of_stockout TEXT,
            notes TEXT,
            company TEXT,
            cobertura REAL,
            status TEXT DEFAULT 'crash',
            date_resolved TEXT
        )
    """)
    existing = [row[1] for row in cursor.execute("PRAGMA table_info(stockouts)").fetchall()]
    if "cobertura" not in existing:
        cursor.execute("ALTER TABLE stockouts ADD COLUMN cobertura REAL")
    if "status" not in existing:
        cursor.execute("ALTER TABLE stockouts ADD COLUMN status TEXT DEFAULT 'crash'")
    if "date_resolved" not in existing:
        cursor.execute("ALTER TABLE stockouts ADD COLUMN date_resolved TEXT")
    conn.commit()

_migrate_db()

procedure = input("Que queres hacer?\n1-Actualizar stocks\n2-Alertar faltantes\n3-Ver KPIs de stockouts\n")


def actualizar_stock():
    load_dotenv()
    
    METABASE_URL = os.getenv("URL")
    USERNAME = os.getenv("METABASE_USER")
    PASSWORD = os.getenv("PASSWORD")
    print("Porfavor espera mientras se actualizan los stocks...")
    
    try:
    # Authenticate
        
        session = requests.Session()
        token = session.post(f"{METABASE_URL}/api/session", json={
            "username": USERNAME,
            "password": PASSWORD
        }).json()["id"]

        # Download the report
        response = session.post(
            f"{METABASE_URL}/api/card/201/query/xlsx",
            headers={"X-Metabase-Session": token}
        )
        print("Connectado exitosamente.")

        with open("temp_report.xlsx", "wb") as f:
            f.write(response.content)

        # Load data from the downloaded file
        df = pd.read_excel("temp_report.xlsx")

    except Exception as e:
        print(f"Error on Metabase connection: {e}")
        return

    stock_ciu = df[df["CtroDistrib"].str.contains("CIUDADELA")]
    stock_moreno = df[(df["Empresa"].str.contains("GRUPO L")) & ((df["CtroDistrib"].str.contains("CDM03 - MORENO 3")) | (df["CtroDistrib"].str.contains("CDM02 - MORENO 2")) | (df["CtroDistrib"].str.contains("CD MORENO")))]
    stock_atlantico = df[df["Empresa"].str.contains("SERVICIOS ATLANTICO SA")]

    with pd.ExcelWriter(TARGET_EXCEL, engine="openpyxl", mode="w") as writer:
        stock_ciu.to_excel(writer, sheet_name="STOCK CIUDADELA", index=False)
        stock_moreno.to_excel(writer, sheet_name="STOCK MORENO", index=False)
        stock_atlantico.to_excel(writer, sheet_name="STOCK ATLANTICO", index=False)

    ATLANTICO_PATH = r"Files/Seguimiento stock descartables ATLANTICO.xlsx"
    VO_PATH = r"Files/Seguimiento stock descartables VO.xlsx"

    write_to_excel(ATLANTICO_PATH, "Stock", stock_atlantico, start_row=2, clear=True)
    write_to_excel(VO_PATH, "Stock", stock_moreno, start_row=2, clear=True)
    write_to_excel(VO_PATH, "Stock", stock_ciu, start_row=len(stock_moreno) + 2)

def write_to_excel(target_path, sheet_name, df, start_row=2, clear=False):

    wb = load_workbook(target_path)
    ws = wb[sheet_name]

    if clear:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=15):
            for cell in row:
                cell.value = None

    for i, (_, row) in enumerate(df.iterrows()):
        for j, value in enumerate(row.iloc[:15]):
            ws.cell(row=start_row + i, column=j + 1, value=value)

    wb.save(target_path)

def _sync_stockouts(df, company):
    df = df.copy()
    df[STOCK_COL] = pd.to_numeric(df[STOCK_COL], errors="coerce")
    df = df.dropna(subset=[STOCK_COL])

    open_db = pd.read_sql(
        "SELECT * FROM stockouts WHERE company = ? AND status != 'resolved'",
        conn, params=(company,)
    )
    open_products = set(open_db["product_name"].tolist())
    crashed = df[df[STOCK_COL] < 0.50]
    crashed_products = set(crashed[PRODUCT_COL].tolist())

    # Products that recovered: were open, now >= 50%
    for product in open_products - crashed_products:
        cursor.execute(
            "UPDATE stockouts SET status = 'resolved', date_resolved = ? "
            "WHERE product_name = ? AND company = ? AND status != 'resolved'",
            (str(today), product, company)
        )

    for _, row in crashed.iterrows():
        product = row[PRODUCT_COL]
        cob = row[STOCK_COL]
        new_status = "broke" if cob <= 0 else "crash"
        notes = None if pd.isna(row.get("Estado")) else row["Estado"]

        if product not in open_products:
            # New crash: insert
            cursor.execute(
                "INSERT INTO stockouts (product_name, date_of_stockout, notes, company, cobertura, status) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (product, str(today), notes, company, cob, new_status)
            )
        elif new_status == "broke":
            # Escalate existing crash to broke
            cursor.execute(
                "UPDATE stockouts SET status = 'broke', cobertura = ? "
                "WHERE product_name = ? AND company = ? AND status = 'crash'",
                (cob, product, company)
            )

    conn.commit()


def alertar_faltantes():

    def send_email(html_body, subject):
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = EMAIL_RECEIVER
        mail.Subject = subject
        mail.HTMLBody = html_body
        mail.Send()

    def build_rows_html(low_df):
        rows = ""
        for _, row in low_df.iterrows():
            product = row[PRODUCT_COL]
            cobertura = f"{row[STOCK_COL] * 100:.2f}%"
            estado = "Pendiente" if pd.isna(row.get("Estado")) else row["Estado"]
            broke_style = ' style="background-color:#ffe0e0;"' if row[STOCK_COL] <= 0 else ""
            rows += f"""
            <tr{broke_style}>
                <td style="padding: 8px; border: 1px solid #ddd;">{product}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align:center;">{cobertura}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align:center;">{estado}</td>
            </tr>
        """
        return rows

    VO_STOCK = pd.read_excel(r"Files/Seguimiento stock descartables VO.xlsx", sheet_name="Necesidad por familia")
    VO_STOCK["Cobertura"] = pd.to_numeric(VO_STOCK["Cobertura"], errors="coerce")
    low_stock_VO_df = VO_STOCK[VO_STOCK["Cobertura"] < 0.50].dropna(subset=["Cobertura"])
    rows_html = build_rows_html(low_stock_VO_df)

    html_body = f"""
    <p>Este es un mail automatico, comprar los siguientes productos con poco stock:</p>
    <table style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif;">
        <thead>
            <tr style="background-color: #4472C4; color: white;">
                <th style="padding: 10px; border: 1px solid #ddd;">Producto</th>
                <th style="padding: 10px; border: 1px solid #ddd;">Cobertura</th>
                <th style="padding: 10px; border: 1px solid #ddd;">Estado</th>
            </tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
    </table>
    """

    send_email(html_body, "⚠️ Alerta productos con stock menor a 50% VO")
    _sync_stockouts(VO_STOCK, "VO")

    # Build email ATLANTICO
    ATLANTICO_STOCK = pd.read_excel(r"Files/Seguimiento stock descartables ATLANTICO.xlsx", "Necesidad por familia")
    ATLANTICO_STOCK["Cobertura"] = pd.to_numeric(ATLANTICO_STOCK["Cobertura"], errors="coerce")
    low_stock_ATLANTICO_df = ATLANTICO_STOCK[ATLANTICO_STOCK["Cobertura"] < 0.50]      
    low_stock_ATLANTICO_df = low_stock_ATLANTICO_df.dropna(subset=["Cobertura"])                                     

    rows_html = ""
    for _, row in low_stock_ATLANTICO_df.iterrows():
        product = row[PRODUCT_COL]
        cobertura = f"{row[STOCK_COL] * 100:.2f}%"
        estado = "Pendiente" if pd.isna(row["Estado"]) else row["Estado"]
        rows_html += f"""
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd;">{product}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align:center;">{cobertura}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align:center;">{estado}</td>
            </tr>
        """

    html_body = f"""
        <p>Este es un mail automatico, comprar los siguientes productos con poco stock:</p>
        <table style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif;">
            <thead>
                <tr style="background-color: #4472C4; color: white;">
                    <th style="padding: 10px; border: 1px solid #ddd;">Producto</th>
                    <th style="padding: 10px; border: 1px solid #ddd;">Cobertura</th>
                    <th style="padding: 10px; border: 1px solid #ddd;">Estado</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
        """

    # Send email ATLATNICO
    send_email(html_body, "⚠️ Alerta productos con stock menor a 50% ATLANTICO")
    _sync_stockouts(ATLANTICO_STOCK, "ATLANTICO")


if (procedure == "1"):
    actualizar_stock()
elif (procedure == "2"):
    alertar_faltantes()
elif (procedure == "3"):
    mostrar_kpis()

